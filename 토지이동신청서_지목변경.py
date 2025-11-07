import re
from copy import copy
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# ========== 경로 / 시트 설정 ==========
SRC_PATH       = "./in/지목변경신청서_통합.xlsx"
SRC_SHEET      = "1"
TEMPLATE_PATH  = "./in/토지이동신청서_페이지.xlsx"
SHEET1_NAME    = "1" 
SHEET2_NAME    = "2"  
OUT_DIR        = Path("./out")

DROP_FIRST_ROW = False

# ========== 원본 컬럼 이름 ==========
COL_GROUP    = "동리"
COL_SIGUNGU  = "시군"
COL_EUPMYEON = "읍면"
COL_DONGRI   = "동리"

COL_BF_JIBUN = "이동전지번"
COL_BF_JIMOK = "이동전지목"
COL_BF_AREA  = "이동전면적"
COL_AF_JIBUN = "이동후지번"
COL_AF_JIMOK = "이동후지목"
COL_AF_AREA  = "이동후면적"

# ========== 1쪽 좌표 ==========
ROW_START_1          = 17
ROWS_PER_PAGE_1      = 6
COL_SGG_1 = 1 
COL_EM_1  = 2 
COL_DR_1  = 4  

COL_BF_JB_1 = 6 
COL_BF_JM_1 = 8  
COL_BF_AR_1 = 10 
COL_AF_JB_1 = 11  
COL_AF_JM_1 = 13
COL_AF_AR_1 = 14 
COL_MAX_1   = 16 

# ========== 2쪽(이후 페이지 공통) 좌표 ==========
ROW_START_2     = 3 
ROWS_PER_PAGE_2 = 30
COL_A_2 = 1  
COL_B_2 = 2 
COL_C_2 = 3 
COL_D_2 = 4
COL_E_2 = 5 
COL_F_2 = 6 
COL_G_2 = 7 
COL_H_2 = 8  
COL_I_2 = 9 
COL_MAX_2 = 30

# ========== 서식 기본값 ==========
DEFAULT_FONT = Font(name="돋움", size=10)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")

# ========== 유틸리티 ==========
def norm_empty(x):
    """NaN/None/'nan'/'none'/'-' → ''"""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except Exception:
        pass
    s = str(x).strip()
    if s.lower() in {"nan", "none"} or s == "-":
        return ""
    return s

def build_merged_map(ws):
    """병합 셀 위치 매핑"""
    mp = {}
    for mr in ws.merged_cells.ranges:
        for r in range(mr.min_row, mr.max_row + 1):
            for c in range(mr.min_col, mr.max_col + 1):
                mp[(r, c)] = (mr.min_row, mr.min_col)
    return mp

def get_merged_range(ws, r, c):
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= r <= rng.max_row and rng.min_col <= c <= rng.max_col:
            return rng
    return None

def write_safe(ws, r, c, value, merged_map):
    """병합 셀 안전 쓰기(좌상단 셀에만 값/서식 반영)"""
    tr, tc = merged_map.get((r, c), (r, c))
    rng = get_merged_range(ws, tr, tc)
    if rng:
        tr, tc = rng.min_row, rng.min_col
    cell = ws.cell(row=tr, column=tc)
    cell.value = value
    if value is not None:
        cell.font = DEFAULT_FONT
        cell.alignment = CENTER_ALIGN

def copy_cell_style(src, dst):
    if src.has_style:
        if src.font: dst.font = copy(src.font)
        if src.alignment: dst.alignment = copy(src.alignment)
        if src.border: dst.border = copy(src.border)
        if src.fill: dst.fill = copy(src.fill)
        dst.number_format = src.number_format
        if src.protection: dst.protection = copy(src.protection)

def _get_merged_range(ws, row, col):
    """(row, col)이 속한 병합 범위를 찾는다. 없으면 None."""
    for rng in ws.merged_cells.ranges:
        if rng.min_row <= row <= rng.max_row and rng.min_col <= col <= rng.max_col:
            return rng
    return None

def clear_sheet_contents(ws, row_start, col_start, row_end, col_end):
    """
    병합 셀 안전하게 초기화:
    - 병합 범위에 속한 셀은 좌상단 앵커 셀에만 값을 쓴다.
    - 동일 병합 범위를 중복으로 지우지 않도록 visited로 관리.
    """
    visited = set()
    for r in range(row_start, row_end + 1):
        for c in range(col_start, col_end + 1):
            rng = _get_merged_range(ws, r, c)
            if rng:
                anchor = (rng.min_row, rng.min_col)
                if anchor in visited:
                    continue
                ws.cell(row=anchor[0], column=anchor[1]).value = None
                visited.add(anchor)
            else:
                ws.cell(row=r, column=c).value = None

def safe_set(ws, row, col, value):
    """
    병합 셀에 값을 쓸 때도 안전하게 앵커에만 기록.
    일반 셀이라면 그대로 쓴다.
    """
    rng = _get_merged_range(ws, row, col)
    if rng:
        row, col = rng.min_row, rng.min_col
    ws.cell(row=row, column=col).value = value

def clone_column_widths(ws, col_start=1, col_end=16):
    """열 너비 복제"""
    for c in range(col_start, col_end + 1):
        l = get_column_letter(c)
        dim = ws.column_dimensions.get(l)
        if dim and dim.width:
            ws.column_dimensions[l].width = dim.width

def _jibun_sort_key(txt: str):
    """지번(산/본번-부번) 자연 정렬용 키"""
    s = str(norm_empty(txt))
    m = re.fullmatch(r"(산)?(\d+)(?:-(\d+))?", s)
    if not m:
        return (2, s)
    is_san = 1 if m.group(1) else 0
    main = int(m.group(2))
    sub  = int(m.group(3)) if m.group(3) else 0
    return (is_san, main, sub)

# ========== 기록 함수(1쪽) ==========
def write_row_sheet1(ws, merged_map, r, row):
    # 시군 / 읍면 / 동리
    write_safe(ws, r, COL_SGG_1, norm_empty(row.get(COL_SIGUNGU, "")), merged_map)
    write_safe(ws, r, COL_EM_1,  norm_empty(row.get(COL_EUPMYEON, "")), merged_map)
    write_safe(ws, r, COL_DR_1,  norm_empty(row.get(COL_DONGRI, "")), merged_map)

    # 이동전
    bf_jb = norm_empty(row.get(COL_BF_JIBUN, ""))
    bf_jm = norm_empty(row.get(COL_BF_JIMOK, ""))
    bf_ar = norm_empty(row.get(COL_BF_AREA, ""))

    write_safe(ws, r, COL_BF_JB_1,     bf_jb, merged_map)
    write_safe(ws, r, COL_BF_JB_1 + 1, bf_jb, merged_map)
    write_safe(ws, r, COL_BF_JM_1,     bf_jm, merged_map)
    write_safe(ws, r, COL_BF_JM_1 + 1, bf_jm, merged_map)
    write_safe(ws, r, COL_BF_AR_1,     bf_ar, merged_map)

    # 이동후
    af_jb = norm_empty(row.get(COL_AF_JIBUN, ""))
    af_jm = norm_empty(row.get(COL_AF_JIMOK, ""))
    af_ar = norm_empty(row.get(COL_AF_AREA, ""))

    write_safe(ws, r, COL_AF_JB_1,     af_jb, merged_map)
    write_safe(ws, r, COL_AF_JB_1 + 1, af_jb, merged_map)
    write_safe(ws, r, COL_AF_JM_1,     af_jm, merged_map)
    write_safe(ws, r, COL_AF_AR_1,     af_ar, merged_map)

# ========== 기록 함수(2쪽 이후 공통) ==========
def write_row_sheet2(ws2, merged_map2, r2, row):
    A = norm_empty(row.get(COL_SIGUNGU, ""))   # 시군
    B = norm_empty(row.get(COL_EUPMYEON, ""))  # 읍면
    C = norm_empty(row.get(COL_DONGRI, ""))    # 동리
    D = norm_empty(row.get(COL_BF_JIBUN, ""))  # 이동전 지번
    E = norm_empty(row.get(COL_BF_JIMOK, ""))  # 이동전 지목
    F = norm_empty(row.get(COL_BF_AREA, ""))   # 이동전 면적
    G = norm_empty(row.get(COL_AF_JIBUN, ""))  # 이동후 지번
    H = norm_empty(row.get(COL_AF_JIMOK, ""))  # 이동후 지목
    I = norm_empty(row.get(COL_AF_AREA, ""))   # 이동후 면적

    write_safe(ws2, r2, COL_A_2, A, merged_map2)  # A
    write_safe(ws2, r2, COL_B_2, B, merged_map2)  # B
    write_safe(ws2, r2, COL_C_2, C, merged_map2)  # C
    write_safe(ws2, r2, COL_D_2, D, merged_map2)  # D
    write_safe(ws2, r2, COL_E_2, E, merged_map2)  # E
    write_safe(ws2, r2, COL_F_2, F, merged_map2)  # F
    write_safe(ws2, r2, COL_G_2, G, merged_map2)  # G
    write_safe(ws2, r2, COL_H_2, H, merged_map2)  # H
    write_safe(ws2, r2, COL_I_2, I, merged_map2)  # I

# ========== 시트 확보(2쪽을 템플릿으로 복제) ==========
def get_or_clone_page_sheet(wb, page_index: int):
    """
    page_index: 2 → '2', 3 → '3', ...
    '2' 시트를 템플릿으로 하여 필요한 번호의 시트를 생성/반환
    """
    name = str(page_index)
    if name in wb.sheetnames:
        return wb[name]

    # 템플릿 시트 확보(없으면 1쪽을 복제해서 2쪽 만들고 그걸 템플릿으로 사용)
    if SHEET2_NAME in wb.sheetnames:
        template = wb[SHEET2_NAME]
    else:
        base = wb[SHEET1_NAME] if SHEET1_NAME in wb.sheetnames else wb.active
        template = wb.copy_worksheet(base)
        template.title = SHEET2_NAME

    new_ws = wb.copy_worksheet(template)
    new_ws.title = name
    return new_ws

# ========== 그룹별 엑셀 작성(1쪽 → 2,3,4… 자동 이어쓰기) ==========
def fill_group_across_pages(wb, df_grp_sorted):
    """
    그룹별 데이터 → 1쪽 또는 2쪽 이상 시트에 채우기
    (6행 이하 → 1쪽만 사용하고 그 때만 초기화,
     7행 이상 → 1쪽/2쪽 전부 초기화하지 않음)
    """
    ws1 = wb[SHEET1_NAME] if SHEET1_NAME in wb.sheetnames else wb.active
    clone_column_widths(ws1, 1, COL_MAX_1)
    merged_map1 = build_merged_map(ws1)

    rows_iter = df_grp_sorted.to_dict(orient="records")
    total_rows = len(rows_iter)
    idx = 0

    # 6행 이하: 1쪽만 사용 + 이때만 초기화 수행
    if total_rows <= ROWS_PER_PAGE_1:
        clear_sheet_contents(ws1, ROW_START_1, 1, ROW_START_1 + ROWS_PER_PAGE_1 - 1, COL_MAX_1)
        r1 = ROW_START_1
        while idx < len(rows_iter):
            write_row_sheet1(ws1, merged_map1, r1, rows_iter[idx])
            r1 += 1
            idx += 1
        return  # 끝

    # 7행 이상: 1쪽도 초기화하지 않음, 2쪽부터도 초기화하지 않음
    #    (기존 내용/서식 유지한 채 덮어쓰기만 수행)
    # 2쪽 템플릿 확보
    ws2_template = get_or_clone_page_sheet(wb, 2)
    clone_column_widths(ws2_template, 1, COL_MAX_2)

    page_index = 2
    while idx < len(rows_iter):
        ws = get_or_clone_page_sheet(wb, page_index)
        clone_column_widths(ws, 1, COL_MAX_2)
        merged_map2 = build_merged_map(ws)

        rows_written = 0
        r2 = ROW_START_2
        while idx < len(rows_iter) and rows_written < ROWS_PER_PAGE_2:
            write_row_sheet2(ws, merged_map2, r2, rows_iter[idx])
            r2 += 1
            idx += 1
            rows_written += 1

        page_index += 1
        

# ========== 메인 실행 ==========
def main():
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    df = pd.read_excel(SRC_PATH, sheet_name=SRC_SHEET)
    df.columns = df.columns.astype(str).str.strip().str.replace("\n", "", regex=False)

    if DROP_FIRST_ROW and len(df) > 0:
        df = df.iloc[1:].reset_index(drop=True)

    req = [
        COL_GROUP, COL_SIGUNGU, COL_EUPMYEON, COL_DONGRI,
        COL_BF_JIBUN, COL_BF_JIMOK, COL_BF_AREA,
        COL_AF_JIBUN, COL_AF_JIMOK, COL_AF_AREA
    ]
    missing = [c for c in req if c not in df.columns]
    if missing:
        raise KeyError(f"필수 컬럼 누락: {missing}")

    # 지번 정렬(이동전 기준)
    df = df.copy()
    df["_sortkey"] = df[COL_BF_JIBUN].astype(str).map(_jibun_sort_key)
    df = df.sort_values([COL_GROUP, "_sortkey"], kind="stable").drop(columns=["_sortkey"])

    groups = sorted(df[COL_GROUP].dropna().astype(str).unique().tolist())
    made = 0

    for g in groups:
        g_str = str(g).strip()
        if g_str in ["-", "", "nan", "none"]:
            continue

        df_grp = df[df[COL_GROUP].astype(str).str.strip() == g_str].copy()
        if df_grp.empty:
            continue

        wb = load_workbook(TEMPLATE_PATH)

        # 그룹별로 1→2→3→… 나눠 쓰기
        fill_group_across_pages(wb, df_grp)

        safe_group = re.sub(r'[\\/:*?"<>|]', "_", g_str)
        out_path = OUT_DIR / f"{safe_group}.xlsx"
        wb.save(out_path)
        made += 1

    print(f"✅ 완료: {made}개 그룹 처리 → {OUT_DIR.resolve()}")

if __name__ == "__main__":
    main()


